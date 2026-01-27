import os
import datetime
import csv
from pathlib import Path
from typing import Dict, List, Any
import openpyxl
from PyPDF2 import PdfReader
from PIL import Image


def get_file_basic_info(file_path: str) -> Dict[str, Any]:
    """获取文件基本元信息"""
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"文件不存在：{file_path}")

    stat = file_path.stat()
    return {
        "file_name": file_path.name,
        "file_path": str(file_path.absolute()),
        "file_size": stat.st_size,
        "file_type": file_path.suffix.lstrip(".").lower(),
        "created_time": datetime.datetime.fromtimestamp(stat.st_ctime).isoformat(),
        "modified_time": datetime.datetime.fromtimestamp(stat.st_mtime).isoformat(),
        "file_exists": True,
    }


def get_text_file_details(file_path: str) -> Dict[str, Any]:
    """获取文本/CSV文件详细信息"""
    details = {}
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            lines = f.readlines()
            details["line_count"] = len(lines)
            details["char_count"] = sum(len(line) for line in lines)

        # CSV特殊处理
        if file_path.endswith((".csv", ".tsv")):
            delimiter = "," if file_path.endswith(".csv") else "\t"
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                reader = csv.reader(f, delimiter=delimiter)
                header = next(reader, [])
                details["csv_columns"] = len(header)
                details["csv_header"] = header[:10]  # 只取前10列避免过长
                details["csv_data_rows"] = len(lines) - 1
    except Exception as e:
        details["text_parse_error"] = str(e)

    return details


def get_excel_file_details(file_path: str) -> Dict[str, Any]:
    """获取Excel文件详细信息"""
    details = {}
    if not openpyxl:
        details["error"] = "未安装openpyxl（需执行：pip install openpyxl）"
        return details

    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        details["sheet_names"] = workbook.sheetnames
        details["sheet_count"] = len(workbook.sheetnames)

        sheet_stats = {}
        for sheet_name in workbook.sheetnames[:3]:  # 只处理前3个工作表
            sheet = workbook[sheet_name]
            sheet_stats[sheet_name] = {
                "row_count": sheet.max_row,
                "column_count": sheet.max_column,
                "has_header": bool(sheet[1]) if sheet.max_row >= 1 else False,
            }
        details["sheet_details"] = sheet_stats
        workbook.close()
    except Exception as e:
        details["excel_parse_error"] = str(e)

    return details


def get_pdf_file_details(file_path: str) -> Dict[str, Any]:
    """获取PDF文件详细信息"""
    details = {}
    if not PdfReader:
        details["error"] = "未安装PyPDF2（需执行：pip install PyPDF2）"
        return details

    try:
        reader = PdfReader(file_path)
        details["page_count"] = len(reader.pages)
        details["pdf_version"] = (
            reader.pdf_header.split(" ")[1] if reader.pdf_header else None
        )
        details["author"] = reader.metadata.author if reader.metadata else "未知"
        details["title"] = reader.metadata.title if reader.metadata else "无标题"
    except Exception as e:
        details["pdf_parse_error"] = str(e)

    return details


def get_image_file_details(file_path: str) -> Dict[str, Any]:
    """获取图片文件详细信息"""
    details = {}
    if not Image:
        details["error"] = "未安装pillow（需执行：pip install pillow）"
        return details

    try:
        with Image.open(file_path) as img:
            details["image_size"] = f"{img.width}x{img.height}"
            details["image_mode"] = img.mode
            details["image_format"] = img.format
            details["image_channels"] = len(img.getbands())
    except Exception as e:
        details["image_parse_error"] = str(e)

    return details


def get_file_details(file_path: str) -> Dict[str, Any]:
    """获取文件完整详细信息（整合所有类型）"""
    try:
        details = get_file_basic_info(file_path)
        file_type = details["file_type"]

        # 根据文件类型补充详细信息
        if file_type in ["txt", "csv", "tsv", "json"]:
            details.update(get_text_file_details(file_path))
        elif file_type in ["xlsx", "xls"]:
            details.update(get_excel_file_details(file_path))
        elif file_type == "pdf":
            details.update(get_pdf_file_details(file_path))
        elif file_type in ["png", "jpg", "jpeg", "gif", "bmp"]:
            details.update(get_image_file_details(file_path))

        details["success"] = True
    except Exception as e:
        details = {"file_path": file_path, "success": False, "error": str(e)}

    return details


def get_multiple_files_details(file_paths: List[str]) -> List[Dict[str, Any]]:
    """批量获取文件信息"""
    return [get_file_details(path) for path in file_paths]
